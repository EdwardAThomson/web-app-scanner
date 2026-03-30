"""Report generation for webscan results.

Supports multiple output formats:
- JSON (always written)
- HTML (Jinja2 template, self-contained single file)
- Markdown
- CSV (findings table)
- PDF (generated from HTML via WeasyPrint)
- XLSX (Excel spreadsheet via openpyxl)
"""

import csv
import json
import os
from datetime import datetime
from pathlib import Path

from rich.console import Console
from rich.table import Table

from webscan import __version__
from webscan.models import Finding, ScanResult
from webscan.utils import ensure_output_dir, timestamp_filename

console = Console()

TEMPLATE_DIR = Path(__file__).parent.parent / "templates"


def write_raw_findings(scan_result: ScanResult, output_dir: str) -> str:
    """Write a flat JSON array of every finding from every module.

    This is the complete, unprocessed list before deduplication.
    """
    out_dir = ensure_output_dir(output_dir)
    filepath = os.path.join(out_dir, "findings-raw.json")

    with open(filepath, "w") as f:
        json.dump([f.to_dict() for f in scan_result.all_findings], f, indent=2)

    return filepath


def write_json_report(
    scan_result: ScanResult, output_dir: str,
    deduped_findings: list[Finding] | None = None,
    diff_result=None,
) -> str:
    """Write scan results to a JSON file. Returns the file path.

    Includes raw per-module findings and a top-level deduped section.
    """
    out_dir = ensure_output_dir(output_dir)
    filename = timestamp_filename("webscan", "json")
    filepath = os.path.join(out_dir, filename)

    data = scan_result.to_dict()
    if deduped_findings is not None:
        data["deduped_findings"] = [f.to_dict() for f in deduped_findings]
        data["summary"]["unique_findings"] = len(deduped_findings)
        data["summary"]["duplicates_merged"] = (
            data["summary"]["total_findings"] - len(deduped_findings)
        )

    if diff_result is not None:
        data["diff"] = diff_result.to_dict()
        data["diff"]["summary"] = diff_result.summary()

    with open(filepath, "w") as f:
        json.dump(data, f, indent=2)

    return filepath


def write_html_report(
    scan_result: ScanResult, output_dir: str, checklist_summary: dict | None = None,
    deduped_findings: list[Finding] | None = None, diff_result=None,
) -> str:
    """Write an HTML report using Jinja2. Returns the file path."""
    from jinja2 import Environment, FileSystemLoader

    env = Environment(
        loader=FileSystemLoader(str(TEMPLATE_DIR)),
        autoescape=True,
    )
    template = env.get_template("report.html.j2")

    # Use deduped findings for the report if available
    findings_source = deduped_findings if deduped_findings is not None else scan_result.all_findings

    # Prepare template data
    summary = scan_result.summary()
    all_findings = [f.to_dict() for f in findings_source]

    # Group findings by severity
    findings_by_severity = {}
    for f in all_findings:
        sev = f["severity"]
        findings_by_severity.setdefault(sev, []).append(f)

    # Calculate duration
    duration = ""
    if scan_result.finished_at and scan_result.started_at:
        delta = scan_result.finished_at - scan_result.started_at
        minutes = int(delta.total_seconds() // 60)
        seconds = int(delta.total_seconds() % 60)
        duration = f"{minutes}m {seconds}s" if minutes else f"{seconds}s"

    html = template.render(
        target=scan_result.target,
        started_at=scan_result.started_at.strftime("%Y-%m-%d %H:%M:%S"),
        duration=duration,
        total_findings=summary["total_findings"],
        by_severity=summary["by_severity"],
        module_results=scan_result.module_results,
        checklist=checklist_summary,
        findings_by_severity=findings_by_severity,
        diff=diff_result.to_dict() if diff_result else None,
        version=__version__,
        generated_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    )

    out_dir = ensure_output_dir(output_dir)
    filename = timestamp_filename("webscan", "html")
    filepath = os.path.join(out_dir, filename)

    with open(filepath, "w") as f:
        f.write(html)

    return filepath


def write_markdown_report(
    scan_result: ScanResult, output_dir: str, checklist_summary: dict | None = None,
    deduped_findings: list[Finding] | None = None, diff_result=None,
) -> str:
    """Write a Markdown report. Returns the file path."""
    summary = scan_result.summary()
    findings_source = deduped_findings if deduped_findings is not None else scan_result.all_findings
    lines = []

    lines.append(f"# webscan Security Report")
    lines.append(f"")
    lines.append(f"**Target:** {scan_result.target}  ")
    lines.append(f"**Date:** {scan_result.started_at.strftime('%Y-%m-%d %H:%M:%S')}  ")
    lines.append(f"**Total findings:** {summary['total_findings']}  ")
    lines.append(f"")

    # Severity summary
    lines.append(f"## Summary")
    lines.append(f"")
    lines.append(f"| Severity | Count |")
    lines.append(f"|----------|-------|")
    for sev in ["critical", "high", "medium", "low", "info"]:
        count = summary["by_severity"].get(sev, 0)
        if count > 0:
            lines.append(f"| {sev.upper()} | {count} |")
    lines.append(f"")

    # Module results
    lines.append(f"## Module Results")
    lines.append(f"")
    lines.append(f"| Module | Status | Findings | Duration |")
    lines.append(f"|--------|--------|----------|----------|")
    for mr in scan_result.module_results:
        status = "OK" if mr.success else f"FAIL ({mr.error[:40]})"
        lines.append(f"| {mr.module_name} | {status} | {len(mr.findings)} | {mr.duration_seconds:.1f}s |")
    lines.append(f"")

    # Baseline diff
    if diff_result is not None:
        ds = diff_result.summary()
        lines.append(f"## Baseline Diff")
        lines.append(f"")
        lines.append(f"| Status | Count |")
        lines.append(f"|--------|-------|")
        lines.append(f"| New | {ds['new']} |")
        lines.append(f"| Fixed | {ds['fixed']} |")
        lines.append(f"| Persistent | {ds['persistent']} |")
        lines.append(f"")
        if diff_result.new:
            lines.append(f"### New Findings ({ds['new']})")
            lines.append(f"")
            for f in diff_result.new:
                lines.append(f"- **[{f.severity.value.upper()}]** {f.title} @ {f.location}")
            lines.append(f"")
        if diff_result.fixed:
            lines.append(f"### Fixed Findings ({ds['fixed']})")
            lines.append(f"")
            for f in diff_result.fixed:
                lines.append(f"- ~~[{f.severity.value.upper()}] {f.title} @ {f.location}~~")
            lines.append(f"")

    # Checklist coverage
    if checklist_summary:
        cs = checklist_summary
        lines.append(f"## Checklist Coverage")
        lines.append(f"")
        lines.append(f"**Modules run:** {', '.join(cs['modules_run'])}")
        lines.append(f"")
        if cs.get("modules_not_run"):
            lines.append(f"**Modules not run:** {', '.join(cs['modules_not_run'])}")
            lines.append(f"")
        lines.append(f"**{cs['coverage_percent']}%** of checklist items covered "
                      f"({cs['automated']} automated + {cs['partial']} partial out of {cs['total_items']} active items)")
        lines.append(f"")
        if cs.get("skipped_items"):
            lines.append(f"### Skipped — run more modules ({cs['skipped']} items)")
            lines.append(f"")
            lines.append(f"| ID | Item | Needs modules |")
            lines.append(f"|-----|------|--------------|")
            for item in cs["skipped_items"]:
                mods = ", ".join(item.modules)
                lines.append(f"| {item.id} | {item.title} | {mods} |")
            lines.append(f"")
        if cs.get("no_module_items"):
            lines.append(f"### No module available ({cs['no_module']} items)")
            lines.append(f"")
            for item in cs["no_module_items"]:
                lines.append(f"- **{item.id}** {item.title} — {item.notes}")
            lines.append(f"")
        if cs["manual_items"]:
            lines.append(f"### Manual review required ({cs['manual']} items)")
            lines.append(f"")
            for item in cs["manual_items"]:
                lines.append(f"- **{item.id}** {item.title} (sev {item.severity})")
            lines.append(f"")

    # Findings by severity
    lines.append(f"## Findings")
    lines.append(f"")
    for sev in ["critical", "high", "medium", "low", "info"]:
        sev_findings = [f for f in findings_source if f.severity.value == sev]
        if not sev_findings:
            continue
        lines.append(f"### {sev.upper()} ({len(sev_findings)})")
        lines.append(f"")
        for f in sev_findings:
            lines.append(f"#### {f.title}")
            lines.append(f"")
            lines.append(f"- **Source:** {f.source} | **Category:** {f.category.value}")
            lines.append(f"- **Location:** {f.location}")
            lines.append(f"- **Description:** {f.description}")
            if f.evidence:
                lines.append(f"- **Evidence:** `{f.evidence[:200]}`")
            if f.remediation:
                lines.append(f"- **Remediation:** {f.remediation}")
            if f.reference:
                lines.append(f"- **Reference:** {f.reference}")
            lines.append(f"")

    lines.append(f"---")
    lines.append(f"*Generated by webscan v{__version__} on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}*")

    out_dir = ensure_output_dir(output_dir)
    filename = timestamp_filename("webscan", "md")
    filepath = os.path.join(out_dir, filename)

    with open(filepath, "w") as f:
        f.write("\n".join(lines))

    return filepath


def write_csv_report(
    scan_result: ScanResult, output_dir: str,
    deduped_findings: list[Finding] | None = None, diff_result=None,
) -> str:
    """Write findings as a CSV file. Returns the file path."""
    from webscan.dedup import _dedup_key

    findings_source = deduped_findings if deduped_findings is not None else scan_result.all_findings
    out_dir = ensure_output_dir(output_dir)
    filename = timestamp_filename("webscan", "csv")
    filepath = os.path.join(out_dir, filename)

    # Build diff status lookup
    diff_status: dict[tuple, str] = {}
    if diff_result is not None:
        for f in diff_result.new:
            diff_status[_dedup_key(f)] = "new"
        for f in diff_result.fixed:
            diff_status[_dedup_key(f)] = "fixed"
        for f in diff_result.persistent:
            diff_status[_dedup_key(f)] = "persistent"

    with open(filepath, "w", newline="") as f:
        writer = csv.writer(f)
        headers = [
            "Severity", "Category", "Source", "Title",
            "Description", "Location", "Evidence", "Remediation", "Reference",
        ]
        if diff_result is not None:
            headers.append("Diff Status")
        writer.writerow(headers)

        for finding in findings_source:
            row = [
                finding.severity.value,
                finding.category.value,
                finding.source,
                finding.title,
                finding.description,
                finding.location,
                finding.evidence[:500],
                finding.remediation,
                finding.reference,
            ]
            if diff_result is not None:
                row.append(diff_status.get(_dedup_key(finding), ""))
            writer.writerow(row)

        # Append fixed findings (not in current scan)
        if diff_result is not None:
            for finding in diff_result.fixed:
                row = [
                    finding.severity.value,
                    finding.category.value,
                    finding.source,
                    finding.title,
                    finding.description,
                    finding.location,
                    finding.evidence[:500],
                    finding.remediation,
                    finding.reference,
                    "fixed",
                ]
                writer.writerow(row)

    return filepath


def write_pdf_report(
    scan_result: ScanResult, output_dir: str, checklist_summary: dict | None = None,
    deduped_findings: list[Finding] | None = None, diff_result=None,
) -> str:
    """Write a PDF report by rendering HTML through WeasyPrint. Returns the file path."""
    from weasyprint import HTML as WeasyprintHTML

    # Generate the HTML first (in memory)
    from jinja2 import Environment, FileSystemLoader

    env = Environment(loader=FileSystemLoader(str(TEMPLATE_DIR)), autoescape=True)
    template = env.get_template("report.html.j2")

    findings_source = deduped_findings if deduped_findings is not None else scan_result.all_findings
    summary = scan_result.summary()
    all_findings = [f.to_dict() for f in findings_source]
    findings_by_severity = {}
    for f in all_findings:
        findings_by_severity.setdefault(f["severity"], []).append(f)

    duration = ""
    if scan_result.finished_at and scan_result.started_at:
        delta = scan_result.finished_at - scan_result.started_at
        minutes = int(delta.total_seconds() // 60)
        seconds = int(delta.total_seconds() % 60)
        duration = f"{minutes}m {seconds}s" if minutes else f"{seconds}s"

    html_content = template.render(
        target=scan_result.target,
        started_at=scan_result.started_at.strftime("%Y-%m-%d %H:%M:%S"),
        duration=duration,
        total_findings=summary["total_findings"],
        by_severity=summary["by_severity"],
        module_results=scan_result.module_results,
        checklist=checklist_summary,
        findings_by_severity=findings_by_severity,
        diff=diff_result.to_dict() if diff_result else None,
        version=__version__,
        generated_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    )

    out_dir = ensure_output_dir(output_dir)
    filename = timestamp_filename("webscan", "pdf")
    filepath = os.path.join(out_dir, filename)

    WeasyprintHTML(string=html_content).write_pdf(filepath)

    return filepath


def write_xlsx_report(
    scan_result: ScanResult, output_dir: str,
    deduped_findings: list[Finding] | None = None, diff_result=None,
) -> str:
    """Write findings as an Excel spreadsheet. Returns the file path."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()

    # --- Summary sheet ---
    ws_summary = wb.active
    ws_summary.title = "Summary"
    summary = scan_result.summary()

    ws_summary.append(["webscan Security Report"])
    ws_summary["A1"].font = Font(size=14, bold=True)
    ws_summary.append([])
    ws_summary.append(["Target", scan_result.target])
    ws_summary.append(["Date", scan_result.started_at.strftime("%Y-%m-%d %H:%M:%S")])
    ws_summary.append(["Total Findings", summary["total_findings"]])
    ws_summary.append([])

    # Severity breakdown
    ws_summary.append(["Severity", "Count"])
    ws_summary[f"A{ws_summary.max_row}"].font = Font(bold=True)
    ws_summary[f"B{ws_summary.max_row}"].font = Font(bold=True)
    sev_colors = {
        "critical": "FF4D6A", "high": "FF6B35",
        "medium": "F0B429", "low": "4DABF7", "info": "868E96",
    }
    for sev in ["critical", "high", "medium", "low", "info"]:
        count = summary["by_severity"].get(sev, 0)
        if count > 0:
            ws_summary.append([sev.upper(), count])
            ws_summary[f"A{ws_summary.max_row}"].fill = PatternFill(
                start_color=sev_colors[sev], end_color=sev_colors[sev], fill_type="solid"
            )
    ws_summary.append([])

    # Module results
    ws_summary.append(["Module", "Status", "Findings", "Duration"])
    for cell in ws_summary[ws_summary.max_row]:
        cell.font = Font(bold=True)
    for mr in scan_result.module_results:
        status = "OK" if mr.success else f"FAIL: {mr.error[:40]}"
        ws_summary.append([mr.module_name, status, len(mr.findings), f"{mr.duration_seconds:.1f}s"])

    ws_summary.column_dimensions["A"].width = 25
    ws_summary.column_dimensions["B"].width = 50

    # --- Findings sheet ---
    ws_findings = wb.create_sheet("Findings")
    headers = ["Severity", "Category", "Source", "Title", "Description",
               "Location", "Evidence", "Remediation", "Reference"]
    ws_findings.append(headers)
    for cell in ws_findings[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="2E3348", end_color="2E3348", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")

    findings_source = deduped_findings if deduped_findings is not None else scan_result.all_findings
    for finding in findings_source:
        ws_findings.append([
            finding.severity.value.upper(),
            finding.category.value,
            finding.source,
            finding.title,
            finding.description,
            finding.location,
            finding.evidence[:500],
            finding.remediation,
            finding.reference,
        ])
        # Color-code severity cell
        row = ws_findings.max_row
        sev = finding.severity.value
        if sev in sev_colors:
            ws_findings[f"A{row}"].fill = PatternFill(
                start_color=sev_colors[sev], end_color=sev_colors[sev], fill_type="solid"
            )

    # Auto-width for key columns
    for col, width in [("A", 12), ("B", 18), ("C", 12), ("D", 50),
                       ("E", 60), ("F", 40), ("G", 40), ("H", 50), ("I", 30)]:
        ws_findings.column_dimensions[col].width = width

    # --- Diff sheet (only when baseline comparison was performed) ---
    if diff_result is not None:
        ws_diff = wb.create_sheet("Diff")
        diff_headers = ["Status", "Severity", "Title", "Location", "Source"]
        ws_diff.append(diff_headers)
        for cell in ws_diff[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="2E3348", end_color="2E3348", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")

        status_colors = {
            "NEW": "FF4D6A",
            "FIXED": "40C057",
            "PERSISTENT": "868E96",
        }
        for status_label, finding_list in [
            ("NEW", diff_result.new),
            ("FIXED", diff_result.fixed),
            ("PERSISTENT", diff_result.persistent),
        ]:
            for f in finding_list:
                ws_diff.append([
                    status_label,
                    f.severity.value.upper(),
                    f.title,
                    f.location,
                    f.source,
                ])
                row = ws_diff.max_row
                color = status_colors[status_label]
                ws_diff[f"A{row}"].fill = PatternFill(
                    start_color=color, end_color=color, fill_type="solid"
                )

        for col, width in [("A", 14), ("B", 12), ("C", 50), ("D", 40), ("E", 20)]:
            ws_diff.column_dimensions[col].width = width

    out_dir = ensure_output_dir(output_dir)
    filename = timestamp_filename("webscan", "xlsx")
    filepath = os.path.join(out_dir, filename)
    wb.save(filepath)

    return filepath


def write_reports(
    scan_result: ScanResult,
    output_dir: str,
    formats: list[str],
    checklist_summary: dict | None = None,
    deduped_findings: list[Finding] | None = None,
    diff_result=None,
) -> dict[str, str]:
    """Write reports in the requested formats. Returns {format: filepath}.

    The JSON report includes both raw per-module findings and the deduped
    canonical list.  All other report formats use the deduped list so
    that human-readable output is concise and free of duplicates.
    """
    paths = {}

    # Flat list of all raw findings (before dedup)
    paths["raw"] = write_raw_findings(scan_result, output_dir)

    # JSON: raw module_results + deduped top-level section + diff
    paths["json"] = write_json_report(scan_result, output_dir, deduped_findings, diff_result)

    # All other formats use deduped findings + optional diff
    paths["html"] = write_html_report(scan_result, output_dir, checklist_summary, deduped_findings, diff_result)

    if "md" in formats or "markdown" in formats:
        paths["md"] = write_markdown_report(scan_result, output_dir, checklist_summary, deduped_findings, diff_result)

    if "csv" in formats:
        paths["csv"] = write_csv_report(scan_result, output_dir, deduped_findings, diff_result)

    if "pdf" in formats:
        paths["pdf"] = write_pdf_report(scan_result, output_dir, checklist_summary, deduped_findings, diff_result)

    if "xlsx" in formats or "xls" in formats or "excel" in formats:
        paths["xlsx"] = write_xlsx_report(scan_result, output_dir, deduped_findings, diff_result)

    return paths


def print_summary(scan_result: ScanResult) -> None:
    """Print a summary table of scan results to the console."""
    summary = scan_result.summary()

    console.print()
    console.print(f"[bold]Target:[/bold] {scan_result.target}")
    console.print(f"[bold]Total findings:[/bold] {summary['total_findings']}")
    console.print(f"[bold]Modules run:[/bold] {summary['modules_run']}")
    if summary["modules_failed"] > 0:
        console.print(f"[bold red]Modules failed:[/bold red] {summary['modules_failed']}")
    console.print()

    # Severity breakdown
    if summary["by_severity"]:
        severity_table = Table(title="Findings by Severity")
        severity_table.add_column("Severity", style="bold")
        severity_table.add_column("Count", justify="right")

        severity_colors = {
            "critical": "bold red",
            "high": "red",
            "medium": "yellow",
            "low": "cyan",
            "info": "dim",
        }

        for sev in ["critical", "high", "medium", "low", "info"]:
            count = summary["by_severity"].get(sev, 0)
            if count > 0:
                color = severity_colors.get(sev, "")
                severity_table.add_row(f"[{color}]{sev.upper()}[/{color}]", str(count))

        console.print(severity_table)
        console.print()

    # Module results table
    module_table = Table(title="Module Results")
    module_table.add_column("Module", style="bold")
    module_table.add_column("Status")
    module_table.add_column("Findings", justify="right")
    module_table.add_column("Duration", justify="right")

    for mr in scan_result.module_results:
        status = "[green]OK[/green]" if mr.success else f"[red]FAIL[/red]"
        if not mr.success and mr.error:
            status += f" ({mr.error[:40]})"
        module_table.add_row(
            mr.module_name,
            status,
            str(len(mr.findings)),
            f"{mr.duration_seconds:.1f}s",
        )

    console.print(module_table)
